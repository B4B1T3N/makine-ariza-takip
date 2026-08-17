"""Makine envanteri, bildirim, rapor, yedekleme ve dışa aktarma testleri."""
from __future__ import annotations

import pytest

from app import config
from app.services import (
    backup_service,
    fault_service,
    machine_service,
    notification_service,
    report_service,
)
from app.services.machine_service import MachineError
from app.utils import export


# --- Makine envanteri -----------------------------------------------------
def test_makine_olusturma_ve_listeleme(machines):
    rows = machine_service.list_machines()
    assert len(rows) == 2
    assert rows[0]["open_faults"] == 0


def test_tekrarli_seri_no_reddedilir(machines):
    with pytest.raises(MachineError, match="zaten kayıtlı"):
        machine_service.create_machine("Başka Makine", "SN-1")


def test_bos_makine_adi_reddedilir(app_db):
    with pytest.raises(MachineError, match="adı boş"):
        machine_service.create_machine("   ")


def test_arama_ve_kategori_filtresi(machines):
    assert len(machine_service.list_machines(search="Pres")) == 1
    assert len(machine_service.list_machines(search="Hat B")) == 1
    assert len(machine_service.list_machines(category="Torna")) == 1
    assert machine_service.list_categories() == ["Pres", "Torna"]
    assert machine_service.list_locations() == ["Hat A", "Hat B"]


def test_acik_arizali_makine_pasife_alinamaz(users, machines):
    fault_service.create_fault(
        machines[0], "Arıza", "", config.PRIORITY_LOW, users["operator"].id
    )
    with pytest.raises(MachineError, match="kapanmamış arıza"):
        machine_service.set_active(machines[0], False)


def test_pasif_makineye_ariza_acilamaz(users, machines):
    machine_service.set_active(machines[0], False)
    with pytest.raises(Exception, match="Pasif"):
        fault_service.create_fault(
            machines[0], "Arıza", "", config.PRIORITY_LOW, users["operator"].id
        )


def test_pasif_makine_varsayilan_listede_gizlenir(machines):
    machine_service.set_active(machines[0], False)
    assert len(machine_service.list_machines()) == 1
    assert len(machine_service.list_machines(include_inactive=True)) == 2


def test_makine_istatistikleri(users, machines):
    fault_id = fault_service.create_fault(
        machines[0], "Arıza", "", config.PRIORITY_LOW, users["operator"].id
    )
    fault_service.change_status(
        fault_id, users["technician"].id, config.STATUS_RESOLVED, "Çözüldü"
    )
    stats = machine_service.machine_stats(machines[0])

    assert stats["total_faults"] == 1
    assert stats["open_faults"] == 0
    assert stats["avg_resolution_hours"] is not None
    assert stats["last_fault_at"]


# --- Bildirimler ----------------------------------------------------------
def test_yeni_ariza_teknisyene_bildirilir(users, machines):
    fault_service.create_fault(
        machines[0], "Arıza", "", config.PRIORITY_URGENT, users["operator"].id
    )
    assert notification_service.unread_count(users["technician"].id) >= 1
    # Kaydı açan kişiye kendi kaydı bildirilmez.
    assert notification_service.unread_count(users["operator"].id) == 0


def test_durum_degisikligi_bildirene_gider(users, machines):
    fault_id = fault_service.create_fault(
        machines[0], "Arıza", "", config.PRIORITY_LOW, users["operator"].id
    )
    fault_service.change_status(
        fault_id, users["technician"].id, config.STATUS_IN_PROGRESS, ""
    )
    rows = notification_service.list_for_user(users["operator"].id)
    assert any("durumu güncellendi" in row["title"] for row in rows)


def test_atama_bildirimi(users, machines):
    fault_id = fault_service.create_fault(
        machines[0], "Arıza", "", config.PRIORITY_LOW, users["operator"].id
    )
    notification_service.mark_all_read(users["technician"].id)
    fault_service.assign(fault_id, users["admin"].id, users["technician"].id)

    rows = notification_service.list_for_user(users["technician"].id, unread_only=True)
    assert any("size atandı" in row["title"] for row in rows)


def test_okundu_isaretleme_ve_silme(users, machines):
    fault_service.create_fault(
        machines[0], "Arıza", "", config.PRIORITY_LOW, users["operator"].id
    )
    technician_id = users["technician"].id
    assert notification_service.unread_count(technician_id) > 0

    notification_service.mark_all_read(technician_id)
    assert notification_service.unread_count(technician_id) == 0

    notification_service.delete_all(technician_id)
    assert notification_service.list_for_user(technician_id) == []


# --- Raporlar -------------------------------------------------------------
@pytest.fixture()
def rapor_verisi(users, machines):
    ids = []
    for index in range(4):
        ids.append(fault_service.create_fault(
            machines[index % 2], f"Arıza {index}", "",
            config.PRIORITY_URGENT if index == 0 else config.PRIORITY_MEDIUM,
            users["operator"].id,
        ))
    # İki kaydı çözüme, birini kapanışa götür.
    fault_service.change_status(ids[0], users["technician"].id, config.STATUS_RESOLVED, "Bitti")
    fault_service.change_status(ids[1], users["technician"].id, config.STATUS_RESOLVED, "Bitti")
    fault_service.change_status(ids[1], users["admin"].id, config.STATUS_CLOSED, "Kapandı")
    return ids


def test_ozet(rapor_verisi):
    summary = report_service.summary()
    assert summary["open_total"] == 2
    assert summary["opened_today"] == 4
    assert summary["closed_today"] == 2
    assert summary["machine_count"] == 2
    assert summary["avg_resolution_hours"] is not None


def test_durum_ve_oncelik_dagilimi(rapor_verisi):
    statuses = report_service.status_distribution()
    assert sum(statuses.values()) == 4
    assert statuses[config.STATUS_CLOSED] == 1

    priorities = report_service.priority_distribution(only_active=True)
    assert sum(priorities.values()) == 2


def test_en_cok_arizalanan_makineler(rapor_verisi):
    top = report_service.top_machines(10)
    assert len(top) == 2
    assert top[0]["fault_count"] >= top[1]["fault_count"]
    assert all("name" in row for row in top)


def test_trend_bos_gunleri_de_icerir(rapor_verisi):
    rows = report_service.trend("2026-01-01", "2026-01-10", "gun")
    assert len(rows) == 10
    assert all(set(row) == {"bucket", "opened", "closed"} for row in rows)

    aylik = report_service.trend("2026-01-01", "2026-03-31", "ay")
    assert len(aylik) == 3


def test_cozum_suresi_raporlari(rapor_verisi):
    assert report_service.avg_resolution_hours() is not None
    by_machine = report_service.resolution_by_machine()
    assert by_machine and all("avg_hours" in row for row in by_machine)


def test_personel_yuku(rapor_verisi, users):
    rows = report_service.workload_by_technician()
    names = [row["full_name"] for row in rows]
    assert "Teknisyen Kişi" in names


# --- Dışa aktarma ---------------------------------------------------------
def test_excel_disa_aktarma(tmp_path):
    from openpyxl import load_workbook

    path = export.export_excel(
        tmp_path / "rapor.xlsx",
        ["Makine", "Sayı"],
        [["Pres 1", 3], ["Torna 1", 5]],
    )
    workbook = load_workbook(path)
    sheet = workbook.active
    assert sheet.max_row == 3
    assert sheet.cell(1, 1).value == "Makine"
    assert sheet.cell(2, 2).value == 3


def test_csv_turkce_karakterlerle_yazilir(tmp_path):
    path = export.export_csv(
        tmp_path / "rapor.csv", ["Makine", "Açıklama"], [["Şanzıman", "Çözüldü"]]
    )
    content = path.read_text(encoding="utf-8-sig")
    assert "Şanzıman" in content
    assert "Çözüldü" in content
    assert ";" in content  # Excel-TR uyumlu ayraç


def test_uzantiya_gore_secim(tmp_path):
    csv_path = export.export_auto(tmp_path / "a.csv", ["x"], [[1]])
    xlsx_path = export.export_auto(tmp_path / "a.xlsx", ["x"], [[1]])
    assert csv_path.suffix == ".csv" and csv_path.stat().st_size > 0
    assert xlsx_path.suffix == ".xlsx" and xlsx_path.stat().st_size > 0


# --- Yedekleme ------------------------------------------------------------
def test_yedek_alma(app_db, users, machines, tmp_path):
    target = backup_service.backup_database(tmp_path / "yedek.db")
    assert target.exists() and target.stat().st_size > 0


def test_tam_yedek_zip(app_db, users, machines, tmp_path):
    import zipfile

    archive = backup_service.backup_full(tmp_path / "tam.zip")
    with zipfile.ZipFile(archive) as zf:
        assert "ariza_takip.db" in zf.namelist()


def test_gecersiz_yedek_reddedilir(app_db, tmp_path):
    bad = tmp_path / "sahte.db"
    bad.write_text("bu bir sqlite dosyasi degil")
    with pytest.raises(backup_service.BackupError):
        backup_service.restore_database(bad)


def test_geri_yukleme_veriyi_dondurur(app_db, users, machines, tmp_path):
    fault_service.create_fault(
        machines[0], "İlk arıza", "", config.PRIORITY_LOW, users["operator"].id
    )
    backup = backup_service.backup_database(tmp_path / "yedek.db")

    fault_service.create_fault(
        machines[0], "Sonraki arıza", "", config.PRIORITY_LOW, users["operator"].id
    )
    assert app_db.scalar("SELECT COUNT(*) FROM faults") == 2

    backup_service.restore_database(backup)
    assert app_db.scalar("SELECT COUNT(*) FROM faults") == 1
