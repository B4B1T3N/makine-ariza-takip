"""Raporlama ekranı: trend grafiği, en çok arızalanan makineler, çözüm süreleri."""
from __future__ import annotations

from datetime import date

from PyQt6.QtCore import QDate
from PyQt6.QtWidgets import (
    QComboBox,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QTableWidget,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)

from app import config
from app.services import report_service
from app.services.auth_service import CurrentUser
from app.ui import style
from app.ui.widgets import common
from app.ui.widgets.charts import ChartCanvas
from app.utils import export
from app.utils.helpers import humanize_duration

PRESETS = [
    ("Son 7 gün", 7),
    ("Son 30 gün", 30),
    ("Son 90 gün", 90),
    ("Son 6 ay", 180),
    ("Son 1 yıl", 365),
    ("Özel aralık", 0),
]

TOP_COLUMNS = ["Makine", "Seri No", "Konum", "Toplam Arıza", "Açık", "Ort. Çözüm Süresi"]
RESOLUTION_COLUMNS = [
    "Makine", "Konum", "Toplam", "Çözülen", "Ortalama", "En Hızlı", "En Yavaş",
]
WORKLOAD_COLUMNS = ["Personel", "Rol", "Açık Kayıt", "Toplam Kayıt", "Ort. Çözüm Süresi"]


class ReportsView(QWidget):
    """Tarih aralığına göre filtrelenebilir raporlar."""

    def __init__(self, user: CurrentUser, parent: QWidget | None = None):
        super().__init__(parent)
        self.user = user
        self._top_rows: list = []
        self._resolution_rows: list = []
        self._workload_rows: list = []
        self._trend_rows: list = []
        self._build()
        self.refresh()

    # --- Arayüz -----------------------------------------------------------
    def _build(self) -> None:
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        container = QWidget()
        scroll.setWidget(container)
        outer.addWidget(scroll)

        root = QVBoxLayout(container)
        root.setContentsMargins(20, 18, 20, 18)
        root.setSpacing(14)

        header_row = QHBoxLayout()
        header_row.addWidget(common.page_header(
            "Raporlar",
            "Arıza trendleri, makine performansı ve çözüm süreleri",
        ))
        header_row.addStretch()
        export_button = QPushButton("Tüm Raporları Excel'e Aktar")
        export_button.setObjectName("Primary")
        export_button.setMinimumHeight(36)
        export_button.clicked.connect(self._export_all)
        header_row.addWidget(export_button)
        root.addLayout(header_row)

        root.addWidget(self._build_filters())

        # --- Özet kartları ---
        cards = QHBoxLayout()
        cards.setSpacing(10)
        self.total_card = common.StatCard("Dönemdeki toplam kayıt", "0", style.PRIMARY)
        self.resolved_card = common.StatCard("Çözülen kayıt", "0", "#0ca30c")
        self.avg_card = common.StatCard("Ortalama çözüm süresi", "-", style.ACCENT)
        self.machines_card = common.StatCard("Arıza veren makine", "0", "#eb6834")
        for card in (self.total_card, self.resolved_card, self.avg_card, self.machines_card):
            cards.addWidget(card)
        root.addLayout(cards)

        # --- Trend grafiği ---
        trend_card = common.Card("Arıza Trendi")
        trend_header = QHBoxLayout()
        trend_hint = QLabel(
            "Seçilen dönemde açılan ve çözüme kavuşan kayıt sayısı"
        )
        trend_hint.setStyleSheet(f"color: {style.TEXT_MUTED}; font-size: 9pt;")
        trend_header.addWidget(trend_hint)
        trend_header.addStretch()

        trend_header.addWidget(QLabel("Gruplama:"))
        self.group_combo = QComboBox()
        self.group_combo.addItem("Günlük", "gun")
        self.group_combo.addItem("Haftalık", "hafta")
        self.group_combo.addItem("Aylık", "ay")
        self.group_combo.setCurrentIndex(0)
        self.group_combo.currentIndexChanged.connect(self.refresh)
        trend_header.addWidget(self.group_combo)
        trend_card.body().addLayout(trend_header)

        self.trend_chart = ChartCanvas(height=3.2)
        trend_card.add(self.trend_chart)
        root.addWidget(trend_card)

        # --- En çok arızalanan makineler ---
        top_card = common.Card("En Çok Arızalanan 10 Makine")
        self.top_chart = ChartCanvas(height=4.2)
        top_card.add(self.top_chart)
        root.addWidget(top_card)

        # --- Tablolar ---
        self.tabs = QTabWidget()

        self.top_table = QTableWidget(0, len(TOP_COLUMNS))
        self.top_table.setHorizontalHeaderLabels(TOP_COLUMNS)
        common.setup_table(self.top_table, stretch_column=0)
        self.tabs.addTab(self.top_table, "Makine Arıza Sayıları")

        self.resolution_table = QTableWidget(0, len(RESOLUTION_COLUMNS))
        self.resolution_table.setHorizontalHeaderLabels(RESOLUTION_COLUMNS)
        common.setup_table(self.resolution_table, stretch_column=0)
        self.tabs.addTab(self.resolution_table, "Çözüm Süreleri")

        self.workload_table = QTableWidget(0, len(WORKLOAD_COLUMNS))
        self.workload_table.setHorizontalHeaderLabels(WORKLOAD_COLUMNS)
        common.setup_table(self.workload_table, stretch_column=0)
        self.tabs.addTab(self.workload_table, "Personel Yükü")

        self.tabs.setMinimumHeight(320)
        root.addWidget(self.tabs)

        footer = QHBoxLayout()
        footer.addStretch()
        table_export = QPushButton("Görüntülenen Tabloyu Aktar")
        table_export.clicked.connect(self._export_current_tab)
        footer.addWidget(table_export)
        root.addLayout(footer)

    def _build_filters(self) -> QWidget:
        from PyQt6.QtWidgets import QDateEdit

        card = common.Card()
        row = QHBoxLayout()
        row.setSpacing(10)

        label = QLabel("Dönem:")
        label.setStyleSheet(f"color: {style.TEXT_MUTED};")
        row.addWidget(label)

        self.preset_combo = QComboBox()
        for text, days in PRESETS:
            self.preset_combo.addItem(text, days)
        self.preset_combo.setCurrentIndex(1)  # Son 30 gün
        self.preset_combo.currentIndexChanged.connect(self._on_preset_changed)
        row.addWidget(self.preset_combo)

        self.date_from = QDateEdit(QDate.currentDate().addDays(-29))
        self.date_from.setCalendarPopup(True)
        self.date_from.setDisplayFormat("dd.MM.yyyy")
        self.date_from.setEnabled(False)
        self.date_from.dateChanged.connect(self._on_custom_date)
        row.addWidget(self.date_from)

        row.addWidget(QLabel("—"))

        self.date_to = QDateEdit(QDate.currentDate())
        self.date_to.setCalendarPopup(True)
        self.date_to.setDisplayFormat("dd.MM.yyyy")
        self.date_to.setEnabled(False)
        self.date_to.dateChanged.connect(self._on_custom_date)
        row.addWidget(self.date_to)

        row.addStretch()
        self.range_label = QLabel()
        self.range_label.setStyleSheet(f"color: {style.TEXT_MUTED}; font-size: 9pt;")
        row.addWidget(self.range_label)

        card.body().addLayout(row)
        return card

    # --- Filtre olayları --------------------------------------------------
    def _on_preset_changed(self) -> None:
        days = self.preset_combo.currentData()
        custom = days == 0
        self.date_from.setEnabled(custom)
        self.date_to.setEnabled(custom)

        if not custom:
            self.date_from.blockSignals(True)
            self.date_to.blockSignals(True)
            self.date_to.setDate(QDate.currentDate())
            self.date_from.setDate(QDate.currentDate().addDays(-(days - 1)))
            self.date_from.blockSignals(False)
            self.date_to.blockSignals(False)

            # Uzun dönemlerde günlük gruplama okunmaz hale gelir.
            if days >= 180:
                self.group_combo.setCurrentIndex(2)
            elif days >= 90:
                self.group_combo.setCurrentIndex(1)
            else:
                self.group_combo.setCurrentIndex(0)
        self.refresh()

    def _on_custom_date(self) -> None:
        if self.preset_combo.currentData() == 0:
            self.refresh()

    def _range(self) -> tuple[str, str]:
        return (
            self.date_from.date().toString("yyyy-MM-dd"),
            self.date_to.date().toString("yyyy-MM-dd"),
        )

    # --- Veri -------------------------------------------------------------
    def refresh(self) -> None:
        date_from, date_to = self._range()
        if date_from > date_to:
            self.range_label.setText("Başlangıç tarihi bitiş tarihinden sonra olamaz.")
            return

        self.range_label.setText(
            f"{self.date_from.date().toString('dd.MM.yyyy')} – "
            f"{self.date_to.date().toString('dd.MM.yyyy')}"
        )

        self._fill_summary(date_from, date_to)
        self._fill_trend(date_from, date_to)
        self._fill_top(date_from, date_to)
        self._fill_resolution(date_from, date_to)
        self._fill_workload()

    def _fill_summary(self, date_from: str, date_to: str) -> None:
        statuses = report_service.status_distribution(date_from, date_to)
        total = sum(statuses.values())
        resolved = statuses[config.STATUS_RESOLVED] + statuses[config.STATUS_CLOSED]
        avg = report_service.avg_resolution_hours(date_from=date_from, date_to=date_to)

        self.total_card.set_value(total)
        self.resolved_card.set_value(
            resolved, f"%{resolved * 100 // total} çözüm oranı" if total else ""
        )
        self.avg_card.set_value(humanize_duration(avg))
        self.machines_card.set_value(len(self._top_rows) if self._top_rows else 0)

    def _fill_trend(self, date_from: str, date_to: str) -> None:
        group = self.group_combo.currentData()
        self._trend_rows = report_service.trend(date_from, date_to, group)

        labels = [_pretty_bucket(row["bucket"], group) for row in self._trend_rows]
        self.trend_chart.draw_trend(
            labels,
            [
                ("Açılan", [row["opened"] for row in self._trend_rows]),
                ("Çözülen", [row["closed"] for row in self._trend_rows]),
            ],
        )

    def _fill_top(self, date_from: str, date_to: str) -> None:
        self._top_rows = report_service.top_machines(10, date_from, date_to)

        if self._top_rows:
            self.top_chart.draw_hbar(
                [row["name"] for row in self._top_rows],
                [row["fault_count"] for row in self._top_rows],
                colors=style.CHART_SINGLE,
                xlabel="Arıza kaydı sayısı",
            )
        else:
            self.top_chart.show_empty("Seçilen dönemde arıza kaydı yok")

        self.top_table.setSortingEnabled(False)
        self.top_table.setRowCount(len(self._top_rows))
        for index, row in enumerate(self._top_rows):
            self.top_table.setItem(index, 0, common.text_item(row["name"]))
            self.top_table.setItem(index, 1, common.text_item(row["serial_no"] or "-"))
            self.top_table.setItem(index, 2, common.text_item(row["location"] or "-"))
            self.top_table.setItem(
                index, 3,
                common.SortableItem(str(row["fault_count"]), row["fault_count"]),
            )
            self.top_table.setItem(
                index, 4,
                common.SortableItem(str(row["open_count"] or 0), row["open_count"] or 0),
            )
            self.top_table.setItem(
                index, 5,
                common.SortableItem(
                    humanize_duration(row["avg_hours"]),
                    row["avg_hours"] if row["avg_hours"] is not None else -1,
                ),
            )
        self.top_table.setSortingEnabled(True)

        # Özet kartı makine sayısını trend ile birlikte günceller.
        self.machines_card.set_value(len(self._top_rows))

    def _fill_resolution(self, date_from: str, date_to: str) -> None:
        self._resolution_rows = report_service.resolution_by_machine(date_from, date_to)

        self.resolution_table.setSortingEnabled(False)
        self.resolution_table.setRowCount(len(self._resolution_rows))
        for index, row in enumerate(self._resolution_rows):
            self.resolution_table.setItem(index, 0, common.text_item(row["name"]))
            self.resolution_table.setItem(index, 1, common.text_item(row["location"] or "-"))
            self.resolution_table.setItem(
                index, 2, common.SortableItem(str(row["total"]), row["total"])
            )
            self.resolution_table.setItem(
                index, 3, common.SortableItem(str(row["resolved"] or 0), row["resolved"] or 0)
            )
            for column, key in ((4, "avg_hours"), (5, "min_hours"), (6, "max_hours")):
                value = row[key]
                self.resolution_table.setItem(
                    index, column,
                    common.SortableItem(
                        humanize_duration(value), value if value is not None else -1
                    ),
                )
        self.resolution_table.setSortingEnabled(True)

    def _fill_workload(self) -> None:
        self._workload_rows = report_service.workload_by_technician()

        self.workload_table.setSortingEnabled(False)
        self.workload_table.setRowCount(len(self._workload_rows))
        for index, row in enumerate(self._workload_rows):
            self.workload_table.setItem(index, 0, common.text_item(row["full_name"]))
            self.workload_table.setItem(
                index, 1, common.text_item(config.ROLE_LABELS.get(row["role"], row["role"]))
            )
            self.workload_table.setItem(
                index, 2,
                common.SortableItem(str(row["open_count"] or 0), row["open_count"] or 0),
            )
            self.workload_table.setItem(
                index, 3,
                common.SortableItem(str(row["total_count"] or 0), row["total_count"] or 0),
            )
            self.workload_table.setItem(
                index, 4,
                common.SortableItem(
                    humanize_duration(row["avg_hours"]),
                    row["avg_hours"] if row["avg_hours"] is not None else -1,
                ),
            )
        self.workload_table.setSortingEnabled(True)

    # --- Dışa aktarma -----------------------------------------------------
    def _current_dataset(self) -> tuple[str, list[str], list[list]]:
        index = self.tabs.currentIndex()
        if index == 0:
            rows = [
                [r["name"], r["serial_no"] or "", r["location"] or "",
                 r["fault_count"], r["open_count"] or 0,
                 _hours(r["avg_hours"])]
                for r in self._top_rows
            ]
            return "Makine Ariza Sayilari", TOP_COLUMNS, rows
        if index == 1:
            rows = [
                [r["name"], r["location"] or "", r["total"], r["resolved"] or 0,
                 _hours(r["avg_hours"]), _hours(r["min_hours"]), _hours(r["max_hours"])]
                for r in self._resolution_rows
            ]
            return "Cozum Sureleri", RESOLUTION_COLUMNS, rows
        rows = [
            [r["full_name"], config.ROLE_LABELS.get(r["role"], r["role"]),
             r["open_count"] or 0, r["total_count"] or 0, _hours(r["avg_hours"])]
            for r in self._workload_rows
        ]
        return "Personel Yuku", WORKLOAD_COLUMNS, rows

    def _export_current_tab(self) -> None:
        title, headers, rows = self._current_dataset()
        if not rows:
            QMessageBox.information(self, "Veri yok", "Aktarılacak veri bulunmuyor.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Raporu kaydet",
            f"{title.lower().replace(' ', '_')}_{date.today().isoformat()}.xlsx",
            export.EXPORT_FILTER,
        )
        if not path:
            return

        try:
            saved = export.export_auto(path, headers, rows, title)
        except export.ExportError as exc:
            QMessageBox.warning(self, "Dışa aktarılamadı", str(exc))
            return
        QMessageBox.information(self, "Dışa aktarıldı", f"Rapor kaydedildi:\n{saved}")

    def _export_all(self) -> None:
        """Tüm rapor tablolarını tek bir Excel dosyasında ayrı sayfalara yazar."""
        if not self._top_rows and not self._trend_rows:
            QMessageBox.information(self, "Veri yok", "Aktarılacak rapor verisi bulunmuyor.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Rapor setini kaydet",
            f"ariza_raporlari_{date.today().isoformat()}.xlsx",
            "Excel Dosyası (*.xlsx)",
        )
        if not path:
            return

        try:
            self._write_workbook(path)
        except Exception as exc:
            QMessageBox.warning(
                self, "Dışa aktarılamadı",
                f"Rapor dosyası yazılamadı:\n{exc}\n\n"
                "Dosya başka bir programda açıksa kapatıp tekrar deneyin.",
            )
            return

        QMessageBox.information(self, "Dışa aktarıldı", f"Rapor seti kaydedildi:\n{path}")

    def _write_workbook(self, path: str) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        group = self.group_combo.currentData()
        sheets = [
            ("Trend", ["Dönem", "Açılan", "Çözülen"],
             [[_pretty_bucket(r["bucket"], group), r["opened"], r["closed"]]
              for r in self._trend_rows]),
        ]
        previous_tab = self.tabs.currentIndex()
        try:
            for index in range(3):
                self.tabs.setCurrentIndex(index)
                sheets.append(self._current_dataset())
        finally:
            self.tabs.setCurrentIndex(previous_tab)

        workbook = Workbook()
        workbook.remove(workbook.active)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2C3E50")

        for title, headers, rows in sheets:
            sheet = workbook.create_sheet(title[:31])
            sheet.append(list(headers))
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            widths = [len(str(h)) for h in headers]
            for row in rows:
                sheet.append(row)
                for i, value in enumerate(row):
                    if i < len(widths):
                        widths[i] = max(widths[i], min(len(str(value)), 50))
            for i, width in enumerate(widths, start=1):
                sheet.column_dimensions[get_column_letter(i)].width = width + 3
            sheet.freeze_panes = "A2"

        info = workbook.create_sheet("Rapor Bilgisi", 0)
        info.append(["Rapor Aralığı", self.range_label.text()])
        info.append(["Oluşturan", self.user.full_name])
        info.append(["Oluşturma Tarihi", date.today().strftime(config.DATE_FMT)])
        info.column_dimensions["A"].width = 20
        info.column_dimensions["B"].width = 34

        workbook.save(path)


def _hours(value: float | None) -> float | str:
    return round(value, 1) if value is not None else ""


def _pretty_bucket(bucket: str, group: str) -> str:
    """SQL bucket anahtarını okunur Türkçe etikete çevirir."""
    try:
        if group == "ay":
            year, month = bucket.split("-")
            months = ["Oca", "Şub", "Mar", "Nis", "May", "Haz",
                      "Tem", "Ağu", "Eyl", "Eki", "Kas", "Ara"]
            return f"{months[int(month) - 1]} {year}"
        if group == "hafta":
            year, week = bucket.split("-W")
            return f"{int(week)}. hafta {year}"
        year, month, day = bucket.split("-")
        return f"{day}.{month}"
    except (ValueError, IndexError):
        return bucket
