"""Tablo verilerini CSV / Excel olarak dışa aktarma."""
from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Iterable, Sequence


class ExportError(Exception):
    """Dışa aktarma hatası (kullanıcıya gösterilebilir mesaj)."""


def export_csv(
    path: str | Path,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
) -> Path:
    """UTF-8 BOM + noktalı virgül ayraç: Excel'in Türkçe yerelinde doğru açılır."""
    target = Path(path)
    try:
        with open(target, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.writer(fh, delimiter=";")
            writer.writerow(headers)
            for row in rows:
                writer.writerow(["" if v is None else v for v in row])
    except OSError as exc:
        raise ExportError(f"CSV yazılamadı: {exc}") from exc
    return target


def export_excel(
    path: str | Path,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    sheet_title: str = "Rapor",
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise ExportError(
            "Excel dışa aktarımı için 'openpyxl' paketi gerekli.\n"
            "Kurulum: pip install openpyxl"
        ) from exc

    target = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or "Rapor"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2C3E50")

    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [len(str(h)) for h in headers]
    for row in rows:
        values = ["" if v is None else v for v in row]
        ws.append(values)
        for i, value in enumerate(values):
            if i < len(widths):
                widths[i] = max(widths[i], min(len(str(value)), 60))

    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width + 3

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    try:
        wb.save(target)
    except OSError as exc:
        raise ExportError(
            f"Excel dosyası yazılamadı: {exc}\n"
            "Dosya başka bir programda açıksa kapatıp tekrar deneyin."
        ) from exc
    return target


def export_auto(
    path: str | Path,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    sheet_title: str = "Rapor",
) -> Path:
    """Uzantıya göre CSV veya Excel'e yazar."""
    target = Path(path)
    if target.suffix.lower() in (".xlsx", ".xlsm"):
        return export_excel(target, headers, rows, sheet_title)
    return export_csv(target, headers, rows)


EXPORT_FILTER = "Excel Dosyası (*.xlsx);;CSV Dosyası (*.csv)"
